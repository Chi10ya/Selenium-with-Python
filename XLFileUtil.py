# ***************  26: Data driven testing using Microsoft Excel + OpenPyXL module
# XLFileUtil.py


import openpyxl

# def getWorkBookSheetName(file, sheetName):
#     workBook = openpyxl.load_workbook(file)
#     sheet = workBook.get_sheet_by_name(sheetName)
#     return sheet
#
# def getRowCount(file, sheetName):
#     sheet = getWorkBookSheetName(file, sheetName)
#     return(sheet.max_row)


def getRowCount(file, sheetName):
    workBook = openpyxl.load_workbook(file)
    sheet = workBook.get_sheet_by_name(sheetName)
    return(sheet.max_row)

def getColumnCount(file, sheetName):
    workBook = openpyxl.load_workbook(file)
    sheet = workBook.get_sheet_by_name(sheetName)
    return (sheet.max_column)

def readDataFile(file, sheetName, rowNum, columnNum):
    workBook = openpyxl.load_workbook(file)
    sheet = workBook.get_sheet_by_name(sheetName)
    return sheet.cell(row=rowNum, column=columnNum).value

def writeDataFile(file, sheetName, rowNum, columnNum, data):
    workBook = openpyxl.load_workbook(file)
    sheet = workBook.get_sheet_by_name(sheetName)
    sheet.cell(row=rowNum, column=columnNum).value = data
    workBook.save(file)


