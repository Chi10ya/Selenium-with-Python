# 24 : How to read data from MS-Excel using OpenPyXL | Data Driven testing.
# 25 : How to write data into Excel using OpenPyXL | Data Driven testing.
# 26: Data driven testing using Microsoft Excel + OpenPyXL module


import openpyxl     # Tutorial 24 : How to read data from MS-Excel using OpenPyXL | Data Driven testing.
import XLFileUtil   # Tutorial 26 :
from selenium import  webdriver

XLpath_ReadData = "C:\\Users\\chaitanya.mohammad\\PycharmProjects\\SDET_SeleniumWithPython\\EmployeeData.xlsx"
XLPath_WriteData = "C:\\Users\\chaitanya.mohammad\\PycharmProjects\\SDET_SeleniumWithPython\\WriteDataToExcel.xlsx"

# *****************Tut 24: How to read data from Ms-Excel using OpenPyXL | DataDriven Testing
"""
Pre-Requisite for to do Ms-Excel data driven is, you need to install OpenPyXL by going to the location of where the python is installed and by using the command
pip install openpyxl
import openpyxl ==. importing the library openpyxl
"""

workbook = openpyxl.load_workbook(XLpath_ReadData)
sheet = workbook.active
#(or)
#sheet = workbook.get_sheet_by_name("Sheet1")
rows = sheet.max_row
cols = sheet.max_column

print(f"Rows = {rows}")
print(f"Columns = {cols}")

for r in range(1, rows+1):
    for c in range(1, cols+1):
        print(sheet.cell(row=r, column=c).value, end="  ")
    print()


# ******************** 25 : How to write data into Excel using OpenPyXL | Data Driven testing.

workbook = openpyxl.load_workbook(XLPath_WriteData)
sheet= workbook.active

for r in range(1, 6):
    for c in range(1, 5):
        if c==1:
            name = "Chaitanya"
        elif c==2:
            name = "Ambica"
        elif c==3:
            name = "Yashvir"
        elif c==4:
            name = "Dhiyanshi"
        sheet.cell(row=r, column=c).value= name
workbook.save(XLPath_WriteData)
workbook.close()


# ***************  26: Data driven testing using Microsoft Excel + OpenPyXL module
# XLFileUtil.py

XLFilePath = "C:\\Users\\chaitanya.mohammad\\PycharmProjects\\SDET_SeleniumWithPython\\26_LoginData.xlsx"
chromeDriverPath = "C:\\Users\\chaitanya.mohammad\\PycharmProjects\\Python_Selenium_BrowserDrivers\\chromedriver.exe"
flightAppURL = "http://newtours.demoaut.com"

driver = webdriver.Chrome(executable_path= chromeDriverPath)
driver.get(flightAppURL)
driver.implicitly_wait(10)
driver.maximize_window()

print (f' ** -->> ** Get Row Count :{XLFileUtil.getRowCount(XLFilePath,"Sheet1")}')
rows = XLFileUtil.getRowCount(XLFilePath, "Sheet1")

for i in range(2, rows+1):
    userName = XLFileUtil.readDataFile(XLFilePath, "Sheet1", r, 1)
    password = XLFileUtil.readDataFile(XLFilePath, "Sheet1", r, 2)

    driver.find_element_by_name("userName").send_keys(userName)
    driver.find_element_by_name("password").send_keys(password)
    driver.find_element_by_name("login").click()

    if driver.title == "Find a Flight: Mercurury Tours:":
        print("Test is passed")
        XLFileUtil.writeDataFile(XLFilePath, "Sheet1", r, 3, "Test Passed")
    else:
        print("Test is failed")
        XLFileUtil.writeDataFile(XLFilePath, "Sheet1", r, 3, "Test Failed")

    driver.find_element_by_link_text("Home").click()















